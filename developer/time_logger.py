from openpyxl import Workbook, load_workbook
import time, os

# excel variables
if os.path.isfile('timelogs.xlsx'):
    workbook = load_workbook('timelogs.xlsx')
else:
    workbook = Workbook()
worksheet = workbook.active

# sets the start time
start_time = time.time()

input(
    "Time is now being tracked. Press enter to close the session and save the log"
)

# sets the end time
end_time = time.time()

# writes to the excel sheet
for row in range(1, 10000):

    if not worksheet['A{}'.format(row)].value:
        worksheet['A{}'.format(row)].value = (end_time - start_time) / 3600

        break

# saves the file
workbook.save('timelogs.xlsx')
